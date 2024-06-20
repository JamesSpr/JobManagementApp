import graphene
from datetime import datetime
from graphql_jwt.decorators import login_required
from django.conf import settings


from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from ..models import Job, Estimate, EstimateHeader, EstimateItem, Invoice, Bill

import environ
env = environ.Env()
environ.Env.read_env()

REPORTS_PATH = f"{env('SHAREPOINT_MAINTENANCE_PATH')}/Admin/Aurify/Reports"
AUD_CURRENCY_FORMAT = "_-[$$-en-AU]* #,##0.00_-;-[$$-en-AU]* #,##0.00_-;_-[$$-en-AU]* -??_-;_-@_-"

def add_headers(sheet, headers, col_offset=0, row_offset=0):
    for i, header in enumerate(headers):
        sheet[f"{chr(65+i+col_offset)}{1+row_offset}"] = header

def add_table(sheet, table_name, table_headers, table_style, col_offset=0, row_offset=0):
    
    # print(sheet, sheet.max_row)
    table = Table(displayName=table_name, 
                    ref=f"{chr(65+col_offset)}{1+row_offset}:{get_column_letter(sheet.max_column)}{str(sheet.max_row - 1)}", 
                    tableColumns=tuple(TableColumn(id=h, name=header) for h, header in enumerate(table_headers, start=1)),
                    tableStyleInfo=table_style,
                    autoFilter=AutoFilter(ref=f"{chr(65+col_offset)}{1+row_offset}:{get_column_letter(sheet.max_column)}{str(sheet.max_row - 1)}"))

    sheet.add_table(table)

def resize_columns(sheet, widths, col_offset=0, row_offset=0):
    for i, width in enumerate(widths):
        sheet.column_dimensions[chr(65+i+col_offset)].width = width + 1

def add_total_row(sheet, total_row, types, col_offset=0, row_offset=0):
    next_row = sheet.max_row + 1
    for i, total in enumerate(total_row):
        sheet[f"{chr(65+i+col_offset)}{next_row}"] = total
        sheet[f"{chr(65+i+col_offset)}{next_row}"].number_format = types[i]
    

def create_table(sheet, table_name, headers, types, col_widths, total_row, style, col_offset=0, row_offset=0):
    add_headers(sheet, headers, col_offset, row_offset)
    add_total_row(sheet, total_row, types, col_offset, row_offset)
    resize_columns(sheet, col_widths, col_offset, row_offset)
    add_table(sheet, table_name, headers, style, col_offset, row_offset)

def get_estimate_breakdowns(est):
    # Calculate estimate costs and profits
    est_costs = 0
    est_profits = 0

    estimate_headers = EstimateHeader.objects.filter(estimate_id=est)
    for header in estimate_headers:
        item = EstimateItem.objects.filter(header_id = header)
        for i in item:
            est_costs += i.extension
            est_profits += i.gross - i.extension

    return est_costs, est_profits

class GenerateFinancialReport(graphene.Mutation):
    class Arguments:
        date_range = graphene.List(graphene.Date)

    success = graphene.Boolean()
    message = graphene.String()

    @classmethod
    @login_required
    def mutate(self, root, info, date_range):
        print(f"Generating Financial Report for timeframe: {date_range[0]} - {date_range[1]}")

        # Create Excel File and Sheets
        report = Workbook()
        report.properties.creator = "Aurify"

        # Data Filters
        approved_estimates = Estimate.objects.filter(job_id__stage="APP", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        underway_estimates = Estimate.objects.filter(job_id__stage="UND", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        invoicing1_estimates = Estimate.objects.filter(job_id__stage="INV", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        invoicing2_estimates = Estimate.objects.filter(job_id__stage="BSA", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        awaiting_payment_estimates = Estimate.objects.filter(job_id__stage="PAY", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        # finalised_estimates = Estimate.objects.filter(job_id__stage="FIN", approval_date__isnull=False, approval_date__gte=date_range[0], approval_date__lt=date_range[1])
        finalised_invoices = Invoice.objects.filter(job_id__stage="FIN", date_paid__isnull=False, date_issued__gte=date_range[0], date_issued__lt=date_range[1])

        now = datetime.now()
        table_style = TableStyleInfo(name="TableStyleLight1", 
                                    showFirstColumn=False,
                                    showLastColumn=False, 
                                    showRowStripes=True)

        ## Approved
        approved = report.create_sheet("Approved")
        approved_table_name = "Approved"
        approved_headers = ["Client", "Job Number", "Location", "Building", "Title", "Approval Date", "Approved Price", "Estimated Costs", "Estimated Profits"]
        approved_types = ["General", "General", "General", "General", "General", "dd/mm/yyyy", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT]
        approved_total_row = ["Total", f"=COUNTA({approved_table_name}[{approved_headers[1]}])", "", "", "", "", f"=SUBTOTAL(109,{approved_table_name}[{approved_headers[6]}])", f"=SUBTOTAL(109,{approved_table_name}[{approved_headers[7]}])", f"=SUBTOTAL(109,{approved_table_name}[{approved_headers[8]}])"]
        approved_col_widths = [len(x) for x in approved_headers]
        
        def approved_values(job, est): 
            costs, profits = get_estimate_breakdowns(est)
            return [
                job.client.get_preferred_name(),
                job.get_identifier(),
                job.location.name,
                job.building,
                job.title,
                est.approval_date,
                est.price,
                costs,
                profits
            ]

        for i, est in enumerate(approved_estimates):
            job = est.job_id
            for val_i, val in enumerate(approved_values(job, est)):
                approved[f"{chr(65 + val_i)}{i+2}"] = val
                approved[f"{chr(65 + val_i)}{i+2}"].number_format = approved_types[val_i]

                if len(str(val)) > approved_col_widths[val_i]:
                    approved_col_widths[val_i] = len(str(val))

        create_table(approved, approved_table_name, approved_headers, approved_types, approved_col_widths, approved_total_row, table_style)

        ## Underway
        underway = report.create_sheet("Underway")
        underway_table_name = "Underway"
        underway_headers = ["Client","Job Number","Location","Building","Title", "Approval Date","Approved Price", "Estimated Costs", "Estimated Profits", "Start Date","Finish Date (EST)"]
        underway_types = ["General", "General", "General", "General", "General", "dd/mm/yyyy", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, "dd/mm/yyyy" ]
        underway_total_row = ["Total", f"=COUNTA({underway_table_name}[{underway_headers[1]}])", "", "", "", "", f"=SUBTOTAL(109,{underway_table_name}[{underway_headers[6]}])", f"=SUBTOTAL(109,{underway_table_name}[{underway_headers[7]}])", f"=SUBTOTAL(109,{underway_table_name}[{underway_headers[8]}])", ""]
        underway_col_widths = [len(x) for x in underway_headers]

        def underway_values(job, est): 
            costs, profits = get_estimate_breakdowns(est)
            return [
                job.client.get_preferred_name(),
                job.get_identifier(),
                job.location.name,
                job.building,
                job.title,
                est.approval_date,
                est.price,
                costs,
                profits,
                job.commencement_date.date()
            ]
        
        for i, est in enumerate(underway_estimates):
            job = est.job_id
            for val_i, val in enumerate(underway_values(job, est)):
                underway[f"{chr(65 + val_i)}{i+2}"] = val
                underway[f"{chr(65 + val_i)}{i+2}"].number_format = underway_types[val_i]
                
                if len(str(val)) > underway_col_widths[val_i]:
                    underway_col_widths[val_i] = len(str(val))

        create_table(underway, underway_table_name, underway_headers, underway_types, underway_col_widths, underway_total_row, table_style)

        ## Invoicing
        invoicing = report.create_sheet("Invoicing")
        invoicing_table_name = "Invoicing"
        invoicing_headers = ["Client","Job Number","Location","Building","Title", "Status", "Approval Date","Approved Price","Estimated Costs", "Estimated Profits", "Close Out Date","Invoice #","Inv Created Date","Days Since Progress"]
        invoicing_types = ["General", "General", "General", "General", "General", "General", "dd/mm/yyyy", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, "dd/mm/yyyy", "General", "dd/mm/yyyy", "General"]
        invoicing_total_row = ["Total", f"=COUNTA({invoicing_table_name}[{invoicing_headers[1]}])", "", "", "", "", "", f"=SUBTOTAL(109,{invoicing_table_name}[{invoicing_headers[7]}])", f"=SUBTOTAL(109,{invoicing_table_name}[{invoicing_headers[8]}])", f"=SUBTOTAL(109,{invoicing_table_name}[{invoicing_headers[9]}])", "", "", "", f"=SUBTOTAL(101,{invoicing_table_name}[{invoicing_headers[13]}])"]
        invoicing_col_widths = [len(x) for x in invoicing_headers]
        
        ### To Submit
        def invoicing1_values(job, est): 
            costs, profits = get_estimate_breakdowns(est)
            return [
                job.client.get_preferred_name(), 
                job.get_identifier(), 
                job.location.name, 
                job.building, 
                job.title, 
                "Invoicing Required", 
                est.approval_date, 
                est.price, 
                costs,
                profits,
                job.close_out_date.date(), 
                "", 
                "", 
                (now.date()-job.close_out_date.date()).days
            ]
        
        for i, est in enumerate(invoicing1_estimates):
            job = est.job_id
            for val_i, val in enumerate(invoicing1_values(job, est)):
                invoicing[f"{chr(65 + val_i)}{i+2}"] = val
                invoicing[f"{chr(65 + val_i)}{i+2}"].number_format = invoicing_types[val_i]
                
                if len(str(val)) > invoicing_col_widths[val_i]:
                    invoicing_col_widths[val_i] = len(str(val))

        ### Pending Submission Approval
        def invoicing2_values(job, est, inv): 
            costs, profits = get_estimate_breakdowns(est)
            return [
                job.client.get_preferred_name(), 
                job.get_identifier(), 
                job.location.name, 
                job.building, 
                job.title,
                "Awaiting Submission Approval",
                est.approval_date,
                est.price,
                costs,
                profits,
                job.close_out_date.date(),
                inv.number,
                inv.date_created,
                (now.date()-inv.date_created).days,
            ]
        
        invoicing2_counter = len(invoicing1_estimates)
        for est in invoicing2_estimates:
            job = est.job_id
            inv = Invoice.objects.filter(job=job)
            for invoice in inv:
                for val_i, val in enumerate(invoicing2_values(job, est, invoice)):
                    invoicing[f"{chr(65 + val_i)}{invoicing2_counter+2}"] = val
                    invoicing[f"{chr(65 + val_i)}{invoicing2_counter+2}"].number_format = invoicing_types[val_i]
                    
                    if len(str(val)) > invoicing_col_widths[val_i]:
                        invoicing_col_widths[val_i] = len(str(val))
                
                invoicing2_counter += 1
                
        create_table(invoicing, invoicing_table_name, invoicing_headers, invoicing_types, invoicing_col_widths, invoicing_total_row, table_style)

        ## Awaiting Payment
        awaiting_payment = report.create_sheet("AwaitingPayment")     
        awaiting_payment_table_name = "AwaitingPayment"
        awaiting_payment_headers = ["Client", "Job Number", "Location", "Building", "Title", "Amount", "Estimated Costs", "Estimated Profits", "Invoice #", "Invoice Sent", "Days Since Submission"]
        awaiting_payment_types = ["General", "General", "General", "General", "General", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, "General", "dd/mm/yyyy", "General"]
        awaiting_payment_total_row = ["Total", f"=COUNTA({awaiting_payment_table_name}[{awaiting_payment_headers[1]}])", "", "", "", f"=SUBTOTAL(109,{awaiting_payment_table_name}[{awaiting_payment_headers[5]}])", f"=SUBTOTAL(109,{awaiting_payment_table_name}[{awaiting_payment_headers[6]}])", f"=SUBTOTAL(109,{awaiting_payment_table_name}[{awaiting_payment_headers[7]}])", "", "", f"=SUBTOTAL(101,{awaiting_payment_table_name}[{awaiting_payment_headers[10]}])"]
        awaiting_payment_col_widths = [len(x) for x in awaiting_payment_headers]

        def awaiting_payment_values(job, est, inv):
            costs, profits = get_estimate_breakdowns(est)
            return [
                job.client.get_preferred_name(),
                job.get_identifier(),
                job.location.name,
                job.building,
                job.title,
                est.price,
                costs,
                profits,
                inv.number,
                inv.date_issued,
                (now.date()-inv.date_issued).days,
            ]
        
        awaiting_payment_counter = 0
        for est in awaiting_payment_estimates:
            job = est.job_id
            inv = Invoice.objects.filter(job=job)
            for invoice in inv:
                for val_i, val in enumerate(awaiting_payment_values(job, est, invoice)):
                    awaiting_payment[f"{chr(65 + val_i)}{awaiting_payment_counter+2}"] = val
                    awaiting_payment[f"{chr(65 + val_i)}{awaiting_payment_counter+2}"].number_format = awaiting_payment_types[val_i]
                    
                    if len(str(val)) > awaiting_payment_col_widths[val_i]:
                        awaiting_payment_col_widths[val_i] = len(str(val))
                
                awaiting_payment_counter += 1

        create_table(awaiting_payment, awaiting_payment_table_name, awaiting_payment_headers, awaiting_payment_types, awaiting_payment_col_widths, awaiting_payment_total_row, table_style)

        ## Finalised
        finalised = report.create_sheet("Finalised")
        finalised_table_name = "Finalised"
        finalised_headers = ["Client", "Job Number", "Location", "Building", "Title", "Amount", "Costs to Date", "Profits to Date", "Invoice #", "Date Invoiced", "Date Paid", "Days to Pay"]
        finalised_types = ["General", "General", "General", "General", "General", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, "General", "dd/mm/yyyy", "dd/mm/yyyy", "General"]
        finalised_total_row = ["Total", f"=COUNTA({finalised_table_name}[{finalised_headers[1]}])", "", "", "", f"=SUBTOTAL(109,{finalised_table_name}[{finalised_headers[5]}])", f"=SUBTOTAL(109,{finalised_table_name}[{finalised_headers[6]}])", f"=SUBTOTAL(109,{finalised_table_name}[{finalised_headers[7]}])", "", "", "", f"=SUBTOTAL(101,{finalised_table_name}[{finalised_headers[11]}])"]
        finalised_col_widths = [len(x) for x in finalised_headers]

        def finalised_values(job, est, inv, bills): 
            days_to_pay = (inv.date_paid-inv.date_issued).days if inv.date_issued else ""
            bill_total = float(sum([b.amount for b in bills])) / 1.1
            return [
                job.client.get_preferred_name(), 
                job.get_identifier(), 
                job.location.name, 
                job.building, 
                job.title,
                est.price,
                bill_total,
                float(est.price) - bill_total, 
                inv.number,
                inv.date_issued,
                inv.date_paid,
                days_to_pay,
            ]
        
        finalised_estimates_counter = 0
        for inv in finalised_invoices:
            job = inv.job
            # inv = Invoice.objects.filter(job=job)
            est = Estimate.objects.get(job_id=job, price=inv.amount)
            bills = Bill.objects.filter(job=job)
            for val_i, val in enumerate(finalised_values(job, est, inv, bills)):
                finalised[f"{chr(65 + val_i)}{finalised_estimates_counter+2}"] = val
                finalised[f"{chr(65 + val_i)}{finalised_estimates_counter+2}"].number_format = finalised_types[val_i]
                
                if len(str(val)) > finalised_col_widths[val_i]:
                    finalised_col_widths[val_i] = len(str(val))
            
            finalised_estimates_counter += 1
        
        create_table(finalised, finalised_table_name, finalised_headers, finalised_types, finalised_col_widths, finalised_total_row, table_style)

        report.remove(report['Sheet'])

        # Summary Sheet
        summary = report.create_sheet("Summary", 0)

        summary_table_name = "Summary"
        summary_headers = ["Job Status", "Total Jobs", "Total Amount", "Estimated Costs", "Estimated Profits"]
        summary_types = ["General", "General", AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT, AUD_CURRENCY_FORMAT]
        summary_total_row = ["Total WIP", f"=SUBTOTAL(109,{summary_table_name}[{summary_headers[1]}])", f"=SUBTOTAL(109,{summary_table_name}[{summary_headers[2]}])", f"=SUBTOTAL(109,{summary_table_name}[{summary_headers[3]}])", f"=SUBTOTAL(109,{summary_table_name}[{summary_headers[4]}])"]
        summary_col_widths = [len(x) for x in summary_headers]

        summary["B6"] = "Approved - Not yet started"
        summary["C6"] = f"={approved_table_name}!B{approved.max_row}"
        summary["D6"] = f"={approved_table_name}!G{approved.max_row}"
        summary["E6"] = f"={approved_table_name}!H{approved.max_row}"
        summary["F6"] = f"={approved_table_name}!I{approved.max_row}"
        summary["B7"] = "Underway"
        summary["C7"] = f"={underway_table_name}!B{underway.max_row}"
        summary["D7"] = f"={underway_table_name}!G{underway.max_row}"
        summary["E7"] = f"={underway_table_name}!H{underway.max_row}"
        summary["F7"] = f"={underway_table_name}!I{underway.max_row}"
        summary["B8"] = "Work Completed - Invoicing"
        summary["C8"] = f"={invoicing_table_name}!B{invoicing.max_row}"
        summary["D8"] = f"={invoicing_table_name}!H{invoicing.max_row}"
        summary["E8"] = f"={invoicing_table_name}!I{invoicing.max_row}"
        summary["F8"] = f"={invoicing_table_name}!J{invoicing.max_row}"
        summary["B9"] = "Invoiced to Client - Awaiting Payment"
        summary["C9"] = f"={awaiting_payment_table_name}!B{awaiting_payment.max_row}"
        summary["D9"] = f"={awaiting_payment_table_name}!F{awaiting_payment.max_row}"
        summary["E9"] = f"={awaiting_payment_table_name}!G{awaiting_payment.max_row}"
        summary["F9"] = f"={awaiting_payment_table_name}!H{awaiting_payment.max_row}"

        for i in range(6, 10):
            summary[f"D{i}"].number_format = AUD_CURRENCY_FORMAT
            summary[f"E{i}"].number_format = AUD_CURRENCY_FORMAT
            summary[f"F{i}"].number_format = AUD_CURRENCY_FORMAT

        create_table(summary, summary_table_name, summary_headers, summary_types, summary_col_widths, summary_total_row, table_style, 1, 4)

        summary["B12"] = "Total Completed & Paid"
        summary["C12"] = f"={finalised_table_name}!B{finalised.max_row}"
        summary["D12"] = f"={finalised_table_name}!F{finalised.max_row}"
        summary["E12"] = f"={finalised_table_name}!G{finalised.max_row}"
        summary["F12"] = f"={finalised_table_name}!H{finalised.max_row}"
        summary["B13"] = "Forecast for EOFY"
        summary["C13"] = f"=C10+C12"
        summary["D13"] = f"=D10+D12"
        summary["E13"] = f"=E10+E12"
        summary["F13"] = f"=F10+F12"

        summary["D12"].number_format = AUD_CURRENCY_FORMAT
        summary["E12"].number_format = AUD_CURRENCY_FORMAT
        summary["F12"].number_format = AUD_CURRENCY_FORMAT
        summary["D13"].number_format = AUD_CURRENCY_FORMAT
        summary["E13"].number_format = AUD_CURRENCY_FORMAT
        summary["F13"].number_format = AUD_CURRENCY_FORMAT

        summary.column_dimensions["B"].width = 32
        summary.column_dimensions["C"].width = 10.82
        summary.column_dimensions["D"].width = 15.73

        summary["B3"] = f"As of {now.strftime('%d/%m/%Y %H:%M')} for the period {date_range[0].strftime('%d/%m/%Y')} - {date_range[1].strftime('%d/%m/%Y')}"
        summary["B2"] = "Aurify Maintenance Financial Summary"
        summary["B2"].font = Font("Aptos", size=16, bold=True, underline='single')

        print(f"Report saved as Financial_{now.strftime('%Y-%m-%d')}.xlsx")
        report.save(f"{REPORTS_PATH}/Financial_{now.strftime('%Y-%m-%d')}.xlsx")

        report.close()

        return self(success=True, message="Successfully Generated Report")
