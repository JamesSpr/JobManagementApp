# Timesheet Script
import os.path
import environ
import requests
import json
from io import BytesIO
from datetime import datetime, timedelta
from django.utils import timezone

from openpyxl import load_workbook

from myob.models import MyobUser
from myob.schema import myob_get
from timesheets.models import Timesheet, Employee, WorkDay, MyobJob
from timesheets.scripts.exchange_email import ExchangeEmail

from exchangelib import FileAttachment, Message, Folder


def get_employee_timesheets_from_email(env, email_account: ExchangeEmail, employees, next_period: datetime):
    """ Reads an email and saves the timesheets for each employee into the next period folder"""

    timesheet_folder = email_account.get_timesheet_folder()
    pay_period_folder_name = f"WE{next_period.day:02d}-{next_period.month:02d}"
    pay_period_folder = folder_exists(timesheet_folder, pay_period_folder_name)

    if not pay_period_folder:
        pay_period_folder = Folder(parent=timesheet_folder, name=pay_period_folder_name)
        pay_period_folder.save()

    
    # Check Windows Folder Structure
    if not os.path.exists(os.path.join(env("SAVE_PATH"), str(datetime.now().year))):
        os.mkdir(os.path.join(env("SAVE_PATH"), str(datetime.now().year)))
    
    save_folder = os.path.join(env("SAVE_PATH"), str(datetime.now().year), pay_period_folder_name)
    if not os.path.exists(save_folder):
        os.mkdir(save_folder)

    timesheets = []

    email_account.account.inbox.refresh()   
    for email in email_account.account.inbox.all():
        if type(email) is not Message:
            continue

        # print(email)

        # Get timesheet attachment from emial
        timesheet_attachment: FileAttachment = None
        for attachment in email.attachments:
            if isinstance(attachment, FileAttachment):
                if attachment.name[-4:-2] == "xl":
                    timesheet_attachment = attachment
    
        if timesheet_attachment is None:
            # Send email notifying no attachment was sent
            print("Timesheet Error")
            # email_account.reply(email, subject="No Timesheet Attached", body="The email you have sent does not have a timesheet attached. Please resend the timesheet to this email account.")
            # email.move(pay_period_folder)
            continue

        # Check timesheet is Valid
        workbook = load_workbook(BytesIO(timesheet_attachment.content), data_only=True)
        sheet = workbook.active
        name = sheet["E4"].value

        if name == "" or name is None:
            print("Timesheet Error")
            # email_account.reply(email, subject="Timesheet Missing Informatio", body="The attached timesheet does not have a name listed. Please correct your timesheet and send back to this email account.") 
            # email.move(pay_period_folder)
            continue

        valid_dates = True
        for i in range(20,6,-1):
            cell_date: datetime = sheet[f"E{i}"].value
            if not cell_date.date() == next_period + timedelta(days=i-20):
                valid_dates = False
            
        if not valid_dates:
            print("Timesheet Error")
            # email_account.reply(email, subject="Timesheet has Incorrect Dates", body="The attached timesheet does not have the correct dates. Please correct your timesheet and send back to this email account.")
            # email.move(pay_period_folder)
            continue

        workbook.close()

        timesheet_path = os.path.join(save_folder, f"{name}.{timesheet_attachment.name.split('.')[-1]}")

        with open(timesheet_path, "wb") as f:
            f.write(timesheet_attachment.content)

        # Add to the timesheets array
        timesheets.append({'employee': email.sender.name, 'email': email.sender.email_address ,'path': timesheet_path})
        
        # Move the email 
        email.move(pay_period_folder)


    return timesheets
    
def folder_exists(parent: Folder, folder_name: str):
    for folder in parent.children:
        if folder.name == folder_name:
            return folder
        
    return False

def get_authenticated_myob_user(myob_env):
    user = MyobUser.objects.get(id=myob_env('MAINTENANCE_UID'))

    if timezone.now() >= (user.access_expires_at - timedelta(minutes=2)):
        
        link = "https://secure.myob.com/oauth2/v1/authorize"
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        payload = {
            'client_id': myob_env('CLIENT_ID'),
            'client_secret': myob_env('CLIENT_SECRET'),
            'refresh_token': user.refresh_token,
            'grant_type':'refresh_token',
        }
        # print(payload)
        response = requests.post(link, data=payload, headers=headers)

        if not response.status_code == 200:
            raise ConnectionError("Bad Connection with MYOB")

        res = json.loads(response.text)

        user.access_token = res['access_token']
        user.refresh_token = res['refresh_token']
        user.access_expires_at = timezone.now() + timedelta(seconds=int(res['expires_in']))
        user.save()
        
        return user
    else:
        return user

def get_pay_period():
    # Get the current or next pay period based off todays date
    latest_timesheet = Timesheet.objects.latest('end_date')
    next_period = latest_timesheet.end_date

    if latest_timesheet.end_date < datetime.now().date():
        two_weeks = timedelta(days=14)
        next_period = latest_timesheet.end_date + two_weeks

    return next_period

def get_myob_employees():
    myob_env = environ.Env()
    environ.Env.read_env("./myob/.env")
    
    myob_user = get_authenticated_myob_user(myob_env)

    endpoint = "Contact/Employee"
    employees = myob_get(myob_user, endpoint, all=True)['Items']

    return employees

def get_timesheet_data(timesheet: dict):
    
    if not os.path.exists(timesheet['path']):
        print("Timesheet not found", timesheet)
        return None

    workbook = load_workbook(filename=timesheet['path'], data_only=True)
    sheet = workbook.active
    name = sheet["E4"].value

    work_type_converter = {
        "Sick Leave": "SICK",
        "Annual Leave": "AL",
        "Public Holiday": "PH",
        "Leave Without Pay": "LWP",
        "Normal Pay": "Normal"
    }

    work_days = []
    for i in range(7,21):
        day = {}
        day.update({'date': sheet[f"E{i}"].value.strftime("%Y-%m-%d")})
        day.update({'hours': sheet[f"J{i}"].value})
        day.update({'job': "" if sheet[f"K{i}"].value is None else sheet[f"K{i}"].value})
        day.update({'work_type': "" if sheet[f"L{i}"].value is None else work_type_converter[sheet[f"L{i}"].value]})
        day.update({'notes': "" if sheet[f"M{i}"].value is None else sheet[f"M{i}"].value})

        work_days.append(day)

    workbook.close()
    
    data = {'name': name, 'work_days': work_days}
    return data

def process_timesheet(data, start_date, end_date):
    start_date = start_date.strftime("%Y-%m-%d")
    end_date = end_date.strftime("%Y-%m-%d")

    employee = Employee.objects.get(name=data['name'])
    if not Timesheet.objects.filter(start_date=start_date, end_date=end_date, employee=employee).exists():
        timesheet = Timesheet()
        timesheet.employee = employee
        timesheet.start_date = start_date
        timesheet.end_date = end_date
        timesheet.save()
    else:
        timesheet = Timesheet.objects.get(start_date=start_date, end_date=end_date, employee=employee)
        # Don't change anything if its been sent to MYOB
        if timesheet.sent_to_myob: return

    for day in data['work_days']:
        work_date = day['date']
        work_day = WorkDay()

        if WorkDay.objects.filter(timesheet = timesheet, date=work_date).exists():
            work_day = WorkDay.objects.get(timesheet = timesheet, date=work_date)

        work_day.timesheet = timesheet
        work_day.date = work_date
        work_day.hours = day['hours']
        work_day.work_type = day['work_type']
        work_day.notes = day['notes']
        
        if " - " in day['job']:
            if MyobJob.objects.filter(number=day['job'].split(" - ")[0]).exists():
                work_day.job = MyobJob.objects.get(number=day['job'].split(" - ")[0])
        elif MyobJob.objects.filter(number=day['job']).exists():
            work_day.job = MyobJob.objects.get(number=day['job'])
        else:
            work_day.job = None

        isWeekend = parse_date(work_date).weekday() >= 5
        allowed_overtime = not(day['work_type'] == "PH" and work_day.job == None) and (employee.pay_basis == "Hourly" or isWeekend)
        work_day.allow_overtime = allowed_overtime

        work_day.save()

def parse_date(date_string):
    return datetime.strptime(date_string, "%Y-%m-%d")

def run():
    """ Connects to an exchange email server and saves relevant timesheet emails"""
    env = environ.Env()
    environ.Env.read_env()

    email = ExchangeEmail()
    email.connect()

    date = datetime.now()
    today = date.strftime("%A")
    if today == "Monday":
        pay_period = get_pay_period() - timedelta(days=14)
    else:
        pay_period = get_pay_period()

    employees = get_myob_employees()
    timesheets = get_employee_timesheets_from_email(env, email, employees, pay_period)

    for timesheet in timesheets:
        data = get_timesheet_data(timesheet)
        process_timesheet(data, pay_period - timedelta(days=13), pay_period)
